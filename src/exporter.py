import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="EDF2F7")
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE
)

COLUMNAS_EXPORT = {
    "fecha": "Fecha",
    "hora": "Hora",
    "tarjeta": "Tarjeta",
    "moneda": "Moneda",
    "monto": "Monto",
    "comercio": "Comercio",
    "estado": "Estado",
    "tipo": "Tipo",
    "categoria": "Categoría",
    "alertas": "Alertas",
}


def _aplicar_estilos(ws, n_cols: int, n_rows: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER

    for row_idx in range(2, n_rows + 2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.border = _CELL_BORDER
            cell.alignment = Alignment(vertical="center")

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, n_rows + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def exportar_excel(df: pd.DataFrame) -> bytes:
    """Genera un archivo Excel con formato y devuelve los bytes para descarga."""
    wb = Workbook()

    # ── Hoja 1: Consumos ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Consumos"

    export_df = df[list(COLUMNAS_EXPORT.keys())].copy()
    export_df.columns = list(COLUMNAS_EXPORT.values())

    for row in dataframe_to_rows(export_df, index=False, header=True):
        ws1.append(row)

    _aplicar_estilos(ws1, len(export_df.columns), len(export_df))

    # ── Hoja 2: Resumen por Categoría ───────────────────────────────────────
    ws2 = wb.create_sheet("Por Categoría")
    cat_df = (
        df[df["moneda"] == "RD"]
        .groupby("categoria")["monto"]
        .agg(Total="sum", Transacciones="count")
        .reset_index()
        .rename(columns={"categoria": "Categoría"})
        .sort_values("Total", ascending=False)
    )
    for row in dataframe_to_rows(cat_df, index=False, header=True):
        ws2.append(row)
    _aplicar_estilos(ws2, len(cat_df.columns), len(cat_df))

    # ── Hoja 3: Resumen por Día ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Día")
    dia_df = (
        df[df["moneda"] == "RD"]
        .groupby("fecha")["monto"]
        .agg(Total="sum", Transacciones="count")
        .reset_index()
        .rename(columns={"fecha": "Fecha"})
        .sort_values("Fecha")
    )
    for row in dataframe_to_rows(dia_df, index=False, header=True):
        ws3.append(row)
    _aplicar_estilos(ws3, len(dia_df.columns), len(dia_df))

    # ── Hoja 4: Alertas ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Alertas")
    alertas_df = df[df["alertas"].notna() & (df["alertas"] != "")][
        list(COLUMNAS_EXPORT.keys())
    ].copy()
    alertas_df.columns = list(COLUMNAS_EXPORT.values())
    for row in dataframe_to_rows(alertas_df, index=False, header=True):
        ws4.append(row)
    if len(alertas_df):
        _aplicar_estilos(ws4, len(alertas_df.columns), len(alertas_df))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
